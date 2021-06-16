import xml.etree.ElementTree as ET
import os
import openpyxl

for file1 in os.listdir("C://Ret_Script"):
    if file1.endswith(".xml"):
        xmlname=ET.parse(file1)
        file2=open(file1,"r")
        str_name=file2.name.split('_')
        mrbts_id=str_name[3].split("-")[1]
        file2.close()
gettroot=xmlname.getroot()
gettroot[0].remove(gettroot[0][0])

wb=openpyxl.load_workbook("RET_Input_Sheet.xlsx")
sheet11=wb["Sheet1"]

max_column=sheet11.max_column

while (sheet11.max_row>1):
    sheet11.delete_rows(2)

row1=2
str_ald=[]
eqm=[]
try:
    for i in range(500):
        mydict = gettroot[0][i].attrib
        if mydict["class"] == "com.nokia.srbts.eqmr:RETU_R":
            #print(mydict["distName"])
            str_ald.append(mydict["distName"].split('/')[3])
            eqm=mydict["distName"].split('/')[0:3]


            for x in gettroot[0][i]:
                sheet11.cell(row=row1, column=4, value=mydict["distName"])
                for col in range(1,max_column+1):
                    cellobjectvalue=sheet11.cell(row=1,column=col)
                    if x.attrib["name"] == cellobjectvalue.value:
                        sheet11.cell(row=row1,column=col,value=x.text)
            row1+=1

except IndexError as e:
    print(e)
#print(eqm)
eqm1=''.join(str(q)+"/" for q in eqm)
#print(eqm1)
row1=2
try:
    for i in range(500):
        mydict1 = gettroot[0][i].attrib
        for y in range(len(str_ald)):
            if mydict1["distName"] == eqm1 + str_ald[y]:
                for x in gettroot[0][i]:
                    sheet11.cell(row=row1,column=7,value=str_ald[y])
                    for col in range(1, max_column + 1):
                        cellobjectvalue = sheet11.cell(row=1, column=col)

                        if x.attrib["name"] == cellobjectvalue.value:
                            sheet11.cell(row=row1, column=col, value=x.text)


                row1 += 1

except IndexError as e:
    pass

row1=2
for k in range(len(str_ald)):
    aldd = eqm1 + str_ald[k]
    try:
        for i in range(500):
            mydict2 = gettroot[0][i].attrib
            if mydict2["class"] == "com.nokia.srbts.eqmr:LOGLINK_R":
                for x in gettroot[0][i]:
                    if x.text == aldd:
                        for l in gettroot[0][i]:
                            sheet11.cell(row=row1, column=3, value=mydict2["distName"])
                            for col in range(1, max_column + 1):
                                cellobjectvalue = sheet11.cell(row=1, column=col)
                                if l.attrib["name"] == cellobjectvalue.value:
                                    sheet11.cell(row=row1, column=col, value=l.text)

                        row1 += 1
    except IndexError as e:
        pass


row1=2
for row2 in range(len(str_ald)):
    temp=sheet11.cell(row=row1,column=3).value
    temp1=temp.split('/')[-1].replace('_R','')
    #print(temp1)
    sheet11.cell(row=row1,column=3,value=temp1)
    row1+=1

row1=2
for row2 in range(len(str_ald)):
    temp=sheet11.cell(row=row1,column=4).value
    temp1=temp.split('/')[-1].replace('_R','')
    #print(temp1)
    sheet11.cell(row=row1,column=4,value=temp1)
    row1+=1

row1=2
for row2 in range(len(str_ald)):
    temp=sheet11.cell(row=row1,column=6).value
    temp1=temp.split('/')[3].replace('_R','')
    #print(temp1)
    sheet11.cell(row=row1,column=6,value=temp1)
    row1+=1

row1=2
for row2 in range(len(str_ald)):
    temp=sheet11.cell(row=row1,column=7).value
    temp1=temp.replace('_R','')
    #print(temp1)
    sheet11.cell(row=row1,column=7,value=temp1)
    row1+=1

row1=2
for row2 in range(len(str_ald)):
    sheet11.cell(row=row1,column=1,value=str_name[4])
    sheet11.cell(row=row1, column=2, value=mrbts_id)
    row1+=1

row1=2
for row2 in range(len(str_ald)):
    #print(sheet11.cell(row=row1,column=10).value)
    if float(sheet11.cell(row=row1,column=10).value) >0:
        sheet11.cell(row=row1, column=10).value=float(sheet11.cell(row=row1, column=10).value) /10.0
    if float(sheet11.cell(row=row1, column=11).value) > 0:
        sheet11.cell(row=row1, column=11).value = float(sheet11.cell(row=row1, column=11).value) / 10.0
    if float(sheet11.cell(row=row1, column=19).value) > 0:
        sheet11.cell(row=row1, column=19).value = float(sheet11.cell(row=row1, column=19).value) / 10.0
    if float(sheet11.cell(row=row1, column=20).value) > 0:
        sheet11.cell(row=row1, column=20).value = float(sheet11.cell(row=row1, column=20).value) / 10.0
    if float(sheet11.cell(row=row1, column=21).value) > 0:
        sheet11.cell(row=row1, column=21).value = float(sheet11.cell(row=row1, column=21).value) / 10.0


    row1+=1
#print(str_ald)

wb.save("RET_Input_Sheet.xlsx")



