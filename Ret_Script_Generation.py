from openpyxl import load_workbook
from yattag import Doc,indent
from datetime import datetime
import os

date = datetime. now(). strftime("%Y_%m_%d-%I_%M_%S_%p")
doc, tag, text, line = Doc().ttl()
date1=datetime.now().strftime("%Y_%m_%d")
if not os.path.exists(date1):
    os.makedirs(date1)

wb=load_workbook("RET_Input_Sheet.xlsx")
sheet11=wb["Sheet1"]
rows=sheet11.max_row
row1=2

doc.asis('<?xml version="1.0" encoding="UTF-8" standalone="no"?>')
doc.asis("<!DOCTYPE raml SYSTEM 'raml20.dtd'>")
with tag('raml', xmlns="raml20.xsd",version="2.0"):
    with tag('cmData', type="plan", scope="all", name="ret_script_10"):
        with tag('header'):
            doc.stag('log', dateTime="2020-03-01", action="created", appInfo="PlanExporter")
        for row2 in range(1, rows):
            with tag('managedObject', klass="LOGLINK", version="EQM20A_2003_002",
                     distName="PLMN-PLMN/MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/HWTOP-1/" + str(
                         sheet11.cell(row1, 3).value), operation="create"):
                line('p', "MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/APEQM-1/" + str(
                    sheet11.cell(row1, 6).value) + "/RSL-1", name="firstEndpointDN")
                line('p',
                     "MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/APEQM-1/" + str(sheet11.cell(row1, 7).value),
                     name="secondEndpointDN")
                line('p', sheet11.cell(row1, 22).value, name="linkModeConfiguration")
                line('p', sheet11.cell(row1, 23).value, name="linkModeDC")
                line('p', sheet11.cell(row1, 24).value, name="linkModeData")
            row1 += 1

        row1 = 2
        for row2 in range(1, rows):
            with tag('managedObject', klass="ALD", version="EQM20A_2003_002",
                     distName="PLMN-PLMN/MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/APEQM-1/" + str(
                         sheet11.cell(row1, 7).value), operation="create"):
                line('p', sheet11.cell(row1, 17).value, name="controlProtocol")
                line('p', sheet11.cell(row1, 18).value, name="productCode")
                line('p', sheet11.cell(row1, 8).value, name="serialNumber")

            row1 += 1

        row1 = 2
        for row2 in range(1, rows):
            with tag('managedObject', klass="RETU", version="EQM20A_2003_002",
                     distName="PLMN-PLMN/MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/APEQM-1/" + str(
                         sheet11.cell(row1, 7).value) + "/" + str(sheet11.cell(row1, 4).value), operation="create"):
                line('p', sheet11.cell(row1, 19).value, name="antBearing")
                line('p', sheet11.cell(row1, 9).value, name="antSerial")
                line('p', sheet11.cell(row1, 12).value, name="baseStationID")
                line('p', sheet11.cell(row1, 20).value, name="maxAngle")
                line('p', sheet11.cell(row1, 11).value, name="mechanicalAngle")
                line('p', sheet11.cell(row1, 21).value, name="minAngle")
                line('p', sheet11.cell(row1, 13).value, name="sectorID")
                line('p', sheet11.cell(row1, 5).value, name="subunitNumber")
                line('p', sheet11.cell(row1, 14).value, name="userNote1")
                line('p', sheet11.cell(row1, 15).value, name="userNote2")
                line('p', sheet11.cell(row1, 10).value, name="angle")
                with tag('list', name="antlDNList"):
                    if sheet11.cell(row1, 16).value == "external":
                        line('p', "external")
                    else:
                        list = []
                        list = sheet11.cell(row1, 16).value.split(',')
                        for i in list:
                            line('p', "MRBTS-" + str(sheet11.cell(row1, 2).value) + "/EQM-1/APEQM-1/" + i)

            row1 += 1

result =indent (
    doc.getvalue(),
    indentation='',
    indent_text=False
)


with open(os.path.join(date1,'Ret_script_'+date+'.xml'), 'w') as f:
    f.write(result)
